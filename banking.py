from difflib import SequenceMatcher
import openpyxl
import sqlite3


def calculate(totalIncome, totalExpense, netIncome, openingBalance, closingBalance):
    for i in range(9, row + 1):
        cell_obj = sheet_obj.cell(row=i, column=7)
        if int(cell_obj.value) > 0:
            totalIncome += cell_obj.value

        else:
            totalExpense += cell_obj.value

    netIncome = totalIncome + totalExpense
    openingBalance = closingBalance - netIncome
    return totalIncome, totalExpense, netIncome, openingBalance

def checkForMatch(payee, list):
    for i in range(0, len(list)):
        result = SequenceMatcher(None, payee, list[i]).ratio()
        if result > 0.7:
            return True, i
    
    return False, -1


# connect to database
try:
    conn = sqlite3.connect('banking.db')
    print("Database connected")
    cursor = conn.cursor()

    table = ''' CREATE TABLE IF NOT EXISTS Banking (
        Payee TEXT NOT NULL,
        Category TEXT NOT NULL,
        Recurring TEXT NOT NULL
        ); '''

    cursor.execute(table)
    print("Table created")
    cursor.close()

except sqlite3.Error as error:
    print("Error while connecting to sqlite", error)

# a list of all payees in the database
payeeList = []
cursor = conn.cursor()
data = cursor.execute("SELECT Payee FROM Banking")
for row in data:
    payeeList.append(row[0])

frequencyDict = {}
recurringCounter = 0

# categories 
food = 0
housing = 0
transportation = 0
shopping = 0
other = 0
entertainment = 0

#keep track of top expenses
prices = []
topExpenses = []

#path to excel file
path = "statement.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row

for i in range(9, row + 1):
    amount = sheet_obj.cell(row=i, column=7)
    if amount.value < 0:
        payee = sheet_obj.cell(row=i, column=5)
        memo = sheet_obj.cell(row=i, column=6)
        
        payeeFull = payee.value + " " + memo.value

        bool, index = checkForMatch(payeeFull, payeeList)

        if bool:
            payee = payeeList[index]
            # frequencyDict[payee] = frequencyDict.get(payee) + 1
            data = cursor.execute(
                "SELECT Category FROM Banking WHERE Payee = ?", (payee,))
            if data.fetchone == "Food":
                food += amount.value
            if data.fetchone == "Housing":
                housing += amount.value
            if data.fetchone == "Transportation":
                transportation += amount.value
            if data.fetchone == "Other":
                other += amount.value
            if data.fetchone == "Entertainment":
                entertainment += amount.value
            if data.fetchone == "Shopping":
                shopping += amount.value
            
            data = cursor.execute("SELECT Recurring FROM Banking WHERE Payee = ?", (payee,))
            if data.fetchone == "Yes":
                recurringCounter += 1
            
            if len(prices) == 0:
                prices.append(amount.value)
                topExpenses.append(payeeFull + str(amount.value))
            else:
                for i in range(0, len(prices)):
                    if amount.value < prices[i]:
                        prices.insert(i, amount.value)
                        topExpenses.insert(i, payeeFull + " " + str(amount.value))
                        break
        else:
            recurringAnswer = input("Is \"" + payeeFull + "\" for $" + str(amount.value) + " a recurring expense?\n1: Yes\n2: No\n")
            categoryAnswer = input("What category does \"" + payeeFull + "\" belong to?\n1: Food\n2: Housing\n3: Transportation\n4: Shopping\n5: Entertainment\n6: Other\n")
            if recurringAnswer == "1":
                recurringAnswer = "Yes"
                recurringCounter += 1
            else:
                recurringAnswer = "No"
            
            if categoryAnswer == "1":
                categoryAnswer = "Food"
                food += amount.value
            if categoryAnswer == "2":
                categoryAnswer = "Housing"
                housing += amount.value
            if categoryAnswer == "3":
                categoryAnswer = "Transportation"
                transportation += amount.value
            if categoryAnswer == "4":
                categoryAnswer = "Shopping"
                shopping += amount.value
            if categoryAnswer == "5":
                categoryAnswer = "Entertainment"
                entertainment += amount.value
            if categoryAnswer == "6":
                categoryAnswer = "Other"
                other += amount.value
            
            cursor.execute("INSERT INTO Banking VALUES (?, ?, ?)", (payeeFull, categoryAnswer, recurringAnswer))
            conn.commit()

            payeeList.append(payeeFull)

            if len(prices) == 0:
                prices.append(amount.value)
                topExpenses.append(payeeFull + " " + str(amount.value))
            else:
                for i in range(0, len(prices)):
                    if amount.value < prices[i]:
                        prices.insert(i, amount.value)
                        topExpenses.insert(i, payeeFull + " " + str(amount.value))
                        break
            
            # frequencyDict[payeeFull] = 1

cursor.close()

# variables
totalIncome = 0
totalExpense = 0
netIncome = 0
openingBalance = 0
closingBalance = 0

currentbalance_obj = sheet_obj.cell(row = 5, column = 1)
currentbalance = currentbalance_obj.value
currentbalance = currentbalance.split(" ")
closingBalance = float(currentbalance[3])

totalIncome, totalExpense, netIncome, openingBalance = calculate(
    totalIncome, totalExpense, netIncome, openingBalance, closingBalance)

print("total income: %.2f" % totalIncome)
print("total expense: %.2f" % totalExpense)
print("net income: %.2f" % netIncome)
print("opening balance: %.2f" % openingBalance)
print("closing balance: %.2f" % closingBalance)
print(payeeList)
print(topExpenses)

